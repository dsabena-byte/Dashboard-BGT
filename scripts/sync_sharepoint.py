#!/usr/bin/env python3
"""
Descarga el Excel publicado en SharePoint via link anonimo y lo convierte
a data.json para el dashboard.

Requiere las variables de entorno:
  SHAREPOINT_URL  -> link "anyone with the link" generado desde SharePoint
                     (ej: https://contoso.sharepoint.com/:x:/s/site/abc?e=xxx)
  SHAREPOINT_SHEET (opcional) -> nombre de la hoja a leer. Si no se setea,
                     el script escanea todas las hojas del workbook y usa
                     la primera donde encuentre los headers esperados.

Headers esperados (case-insensitive, acepta sinonimos):
  presupuesto, nombre cta, año, mes, concepto, $, usd
"""

from __future__ import annotations

import io
import json
import os
import sys
import time
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urlparse, urlunparse, parse_qsl, urlencode

import requests
from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parent.parent
OUTPUT = REPO_ROOT / "data.json"

# Mapeo header_normalizado -> campo interno.
# Notar: "cuenta" sola NO mapea aca porque en algunas planillas "CUENTA"
# es un codigo numerico (ej 6202010072), mientras que la cuenta que el
# dashboard necesita es el NOMBRE de la cuenta (ej PUBLICIDAD EXHIBICIONES).
COLUMN_ALIASES = {
    "presupuesto": "presupuesto",
    "tipo": "presupuesto",
    "tipo de presupuesto": "presupuesto",
    # OJO: "cuenta" sola se mapea a un campo distinto a proposito, porque en
    # planillas tipo "Seguimiento BGT" la columna "CUENTA" es un codigo numerico
    # (ej 6202010072) y el dashboard usa el nombre legible ("NOMBRE CTA").
    "cuenta": "_cuenta_codigo",
    "nombre cta": "cuenta",
    "nombre cuenta": "cuenta",
    "nombre de cuenta": "cuenta",
    "categoria": "cuenta",
    "categoria contable": "cuenta",
    "anio": "anio",
    "ano": "anio",
    "año": "anio",
    "year": "anio",
    "mes": "mes",
    "month": "mes",
    "concepto": "concepto",
    "subcuenta": "concepto",
    "$": "ars",
    "ars": "ars",
    "importe ars": "ars",
    "monto ars": "ars",
    "pesos": "ars",
    "usd": "usd",
    "u$s": "usd",
    "importe usd": "usd",
    "monto usd": "usd",
    "dolares": "usd",
    "dólares": "usd",
}

REQUIRED = ["presupuesto", "cuenta", "anio", "mes", "concepto", "ars", "usd"]

MESES_VALIDOS = {
    "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
    "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE",
}

MIN_MATCH = 5  # cuantas columnas REQUIRED tiene que matchear una fila para tratarla como headers
HEADER_SCAN_ROWS = 25

USER_AGENT = "Mozilla/5.0 (SharePointSync Dashboard-BGT)"
HTTP_TIMEOUT = 60
MAX_RETRIES = 3          # reintentos ante fallos de red transitorios
RETRY_BACKOFF = 2        # segundos base; crece 2s, 4s, 8s...

# Firmas que delatan que SharePoint devolvio una pagina de login / error HTML
# en vez del archivo. Se buscan (en minuscula) dentro de los primeros bytes.
LOGIN_HINTS = (
    "sign in", "iniciar sesi", "login", "_layouts/15/error.aspx",
    "the link you have accessed", "el vinculo al que accedi",
    "this link has expired", "ha expirado", "access denied", "acceso denegado",
)


def build_download_urls(share_url: str) -> list[str]:
    """Devuelve variantes de URL de descarga directa para probar en orden.

    SharePoint/OneDrive acepta forzar la descarga de un share link agregando
    el parametro ?download=1. Igual probamos tambien la URL original por si la
    variante con download rompe la firma del link.
    """
    parsed = urlparse(share_url)
    query = dict(parse_qsl(parsed.query))

    variants: list[str] = []

    # 1) original + download=1 (la que viene funcionando)
    q_dl = dict(query)
    q_dl["download"] = "1"
    variants.append(urlunparse(parsed._replace(query=urlencode(q_dl))))

    # 2) original tal cual (algunos tenants sirven el binario directo)
    variants.append(share_url)

    # Quitar duplicados conservando el orden.
    seen: set[str] = set()
    out: list[str] = []
    for v in variants:
        if v not in seen:
            seen.add(v)
            out.append(v)
    return out


def _looks_like_login(blob: bytes) -> bool:
    head = blob[:4000].decode("utf-8", errors="ignore").lower()
    return any(hint in head for hint in LOGIN_HINTS)


def _fetch(url: str) -> requests.Response:
    """GET con reintentos ante errores de red / 5xx transitorios."""
    last_exc: Exception | None = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = requests.get(
                url,
                headers={"User-Agent": USER_AGENT},
                allow_redirects=True,
                timeout=HTTP_TIMEOUT,
            )
        except requests.RequestException as exc:
            last_exc = exc
            if attempt < MAX_RETRIES:
                wait = RETRY_BACKOFF * (2 ** (attempt - 1))
                print(f"  Fallo de red ({exc.__class__.__name__}), "
                      f"reintentando en {wait}s ({attempt}/{MAX_RETRIES})...")
                time.sleep(wait)
                continue
            raise RuntimeError(
                f"No se pudo conectar a SharePoint tras {MAX_RETRIES} intentos: {exc}"
            ) from exc

        # 5xx -> transitorio, reintentar. 4xx -> permanente, cortar.
        if 500 <= resp.status_code < 600 and attempt < MAX_RETRIES:
            wait = RETRY_BACKOFF * (2 ** (attempt - 1))
            print(f"  HTTP {resp.status_code} del servidor, "
                  f"reintentando en {wait}s ({attempt}/{MAX_RETRIES})...")
            time.sleep(wait)
            continue
        return resp

    # Solo llega aca si se agotaron los reintentos por RequestException.
    raise RuntimeError(f"No se pudo conectar a SharePoint: {last_exc}")


def download_excel(share_url: str) -> bytes:
    urls = build_download_urls(share_url)
    problems: list[str] = []

    for url in urls:
        resp = _fetch(url)

        if resp.status_code != 200:
            problems.append(f"HTTP {resp.status_code} en {url[:80]}...")
            continue

        blob = resp.content
        if blob[:4] == b"PK\x03\x04":
            return blob  # xlsx valido

        ctype = resp.headers.get("Content-Type", "")
        if _looks_like_login(blob):
            problems.append(
                f"{url[:80]}... devolvio una pagina de login/expirado "
                f"(Content-Type: {ctype})"
            )
        else:
            problems.append(
                f"{url[:80]}... no devolvio un .xlsx (Content-Type: {ctype})"
            )

    raise RuntimeError(
        "No se pudo descargar el Excel desde SharePoint. Causa mas probable: "
        "el link caduco o ya no es publico. Regenera el link como "
        "'Cualquier persona con el vinculo' y actualiza el secret SHAREPOINT_URL.\n"
        "Detalle de los intentos:\n  - " + "\n  - ".join(problems)
    )


def normalize_header(h) -> str:
    key = (str(h) if h is not None else "").strip().lower()
    return COLUMN_ALIASES.get(key, key)


def find_header_row(rows: list[tuple]) -> tuple[int, list[str]] | None:
    for i, row in enumerate(rows[:HEADER_SCAN_ROWS]):
        normalized = [normalize_header(c) for c in row]
        matches = sum(1 for c in REQUIRED if c in normalized)
        if matches >= MIN_MATCH:
            return i, normalized
    return None


def parse_sheet(ws, sheet_name: str) -> tuple[list[list] | None, str]:
    """Devuelve (rows, log_msg). rows=None si la hoja no se pudo parsear."""
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return None, f"hoja '{sheet_name}': vacia"

    found = find_header_row(all_rows)
    if not found:
        preview = [
            [("" if c is None else str(c))[:25] for c in r[:8]]
            for r in all_rows[:3]
        ]
        return None, (
            f"hoja '{sheet_name}' ({len(all_rows)} filas): "
            f"no se encontraron headers. Primeras filas: {preview}"
        )

    header_idx, headers = found
    missing = [c for c in REQUIRED if c not in headers]
    if missing:
        return None, (
            f"hoja '{sheet_name}': headers en fila {header_idx+1} pero "
            f"faltan columnas {missing}. Encontradas: {headers}"
        )

    idx = {col: headers.index(col) for col in REQUIRED}
    out: list[list] = []
    skipped = 0
    for row in all_rows[header_idx + 1:]:
        if row is None or all(v is None or v == "" for v in row):
            continue
        try:
            presupuesto = str(row[idx["presupuesto"]] or "").strip()
            cuenta = str(row[idx["cuenta"]] or "").strip()
            anio_raw = row[idx["anio"]]
            anio = str(anio_raw if anio_raw is not None else "").strip().split(".")[0]
            mes = str(row[idx["mes"]] or "").strip().upper()
            concepto = str(row[idx["concepto"]] or "").strip()
            ars_raw = row[idx["ars"]]
            usd_raw = row[idx["usd"]]
        except IndexError:
            skipped += 1
            continue

        if not presupuesto or presupuesto.lower() == "none":
            skipped += 1
            continue
        if mes not in MESES_VALIDOS:
            skipped += 1
            continue
        try:
            ars_num = float(ars_raw) if ars_raw not in (None, "", "-") else 0.0
            usd_num = float(usd_raw) if usd_raw not in (None, "", "-") else 0.0
        except (TypeError, ValueError):
            skipped += 1
            continue

        out.append([
            presupuesto, cuenta, anio, mes, concepto,
            round(ars_num), round(usd_num),
        ])

    msg = (
        f"hoja '{sheet_name}': headers en fila {header_idx+1}, "
        f"{len(out)} filas validas, {skipped} descartadas"
    )
    return out, msg


def parse_workbook(blob: bytes, sheet_name: str | None) -> list[list]:
    wb = load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
    print(f"  Hojas en el workbook: {wb.sheetnames}")

    sheets_to_try = [sheet_name] if sheet_name else wb.sheetnames
    errors = []

    for sn in sheets_to_try:
        if sn not in wb.sheetnames:
            errors.append(f"hoja '{sn}' no existe en el workbook")
            continue
        rows, log = parse_sheet(wb[sn], sn)
        print(f"  -> {log}")
        if rows is not None and len(rows) > 0:
            return rows
        errors.append(log)

    raise RuntimeError(
        "No se pudo extraer datos de ninguna hoja. Detalle:\n  - "
        + "\n  - ".join(errors)
    )


def main() -> int:
    share_url = os.environ.get("SHAREPOINT_URL")
    if not share_url:
        print("ERROR: falta la variable de entorno SHAREPOINT_URL", file=sys.stderr)
        return 1
    sheet = os.environ.get("SHAREPOINT_SHEET") or None

    print("Descargando Excel desde SharePoint...")
    blob = download_excel(share_url)
    print(f"Descarga OK ({len(blob)} bytes). Parseando hoja={sheet or 'auto'}...")
    rows = parse_workbook(blob, sheet)
    print(f"Total: {len(rows)} filas validas.")

    payload = {
        "syncedAt": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "source": "sharepoint",
        "columns": ["presupuesto", "cuenta", "anio", "mes", "concepto", "ars", "usd"],
        "rows": rows,
    }
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=0), encoding="utf-8")
    print(f"Escrito {OUTPUT}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
